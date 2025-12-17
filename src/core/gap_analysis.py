#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
差异分析模块 - 生成SKU×日期的差异汇总表
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class GapAnalyzer:
    """差异分析器"""

    def __init__(self, schedule_file, po_original_file, po_optimized_file):
        """
        初始化差异分析器

        Args:
            schedule_file: 排程目标文件
            po_original_file: 原始PO文件
            po_optimized_file: 优化后PO文件
        """
        self.schedule_df = pd.read_excel(schedule_file)
        self.po_original_df = pd.read_excel(po_original_file)
        self.po_optimized_df = pd.read_excel(po_optimized_file)

        # 标准化PO清单列名（适配新格式）
        column_mapping = {
            'SKU/Spart': 'SKU',
            '发运行数量': '数量',
            '要求交付日期': '修改要货日期'
        }
        self.po_original_df = self.po_original_df.rename(columns=column_mapping)
        self.po_optimized_df = self.po_optimized_df.rename(columns=column_mapping)

    def aggregate_po_by_date(self, po_df, date_column='修改要货日期'):
        """
        按日期汇总PO数量

        Args:
            po_df: PO DataFrame
            date_column: 日期列名

        Returns:
            DataFrame: SKU×日期的汇总表
        """
        # 确保日期列为datetime
        po_df[date_column] = pd.to_datetime(po_df[date_column])

        # 按SKU和日期汇总
        agg_df = po_df.groupby(['SKU', date_column])['数量'].sum().reset_index()
        agg_df = agg_df.rename(columns={date_column: '日期', '数量': 'PO数量'})

        return agg_df

    def create_gap_table(self):
        """
        创建差异汇总表（按周维度）

        Returns:
            DataFrame: SKU×周次的差异表
            dict: 包含各部分数据的字典
        """
        # 1. 为排程目标和PO数据添加week_num（如果不存在）
        schedule_df = self.schedule_df.copy()
        if 'week_num' not in schedule_df.columns:
            schedule_df['week_num'] = schedule_df['日期'].apply(
                lambda x: x.isocalendar()[0] * 100 + x.isocalendar()[1]
            )

        # 2. 按SKU+week_num汇总排程目标
        schedule_weekly = schedule_df.groupby(['SKU', 'week_num'])['计划产量'].sum().reset_index()
        schedule_pivot = schedule_weekly.pivot_table(
            index='SKU',
            columns='week_num',
            values='计划产量',
            fill_value=0
        )

        # 3. 按SKU+week_num汇总PO数据
        po_optimized_df = self.po_optimized_df.copy()
        if 'week_num' not in po_optimized_df.columns:
            po_optimized_df['日期'] = pd.to_datetime(po_optimized_df['修改要货日期'])
            po_optimized_df['week_num'] = po_optimized_df['日期'].apply(
                lambda x: x.isocalendar()[0] * 100 + x.isocalendar()[1]
            )

        po_weekly = po_optimized_df.groupby(['SKU', 'week_num'])['数量'].sum().reset_index()
        po_pivot = po_weekly.pivot_table(
            index='SKU',
            columns='week_num',
            values='数量',
            fill_value=0
        )

        # 4. 确保两个表有相同的周次列
        all_weeks = sorted(set(schedule_pivot.columns) | set(po_pivot.columns))

        # 重建两个表，确保列一致
        schedule_aligned = schedule_pivot.reindex(columns=all_weeks, fill_value=0)
        po_aligned = po_pivot.reindex(columns=all_weeks, fill_value=0)

        # 确保两个表有相同的SKU
        all_skus = sorted(set(schedule_aligned.index) | set(po_aligned.index))
        schedule_aligned = schedule_aligned.reindex(index=all_skus, fill_value=0)
        po_aligned = po_aligned.reindex(index=all_skus, fill_value=0)

        # 5. 计算差异 (排程目标 - PO汇总)
        gap = schedule_aligned - po_aligned

        # 6. 生成周次标签（格式：2025W50）
        week_labels = []
        for week_num in all_weeks:
            year = int(week_num // 100)
            week = int(week_num % 100)
            week_labels.append(f"{year}W{week:02d}")

        return {
            'gap': gap,
            'schedule': schedule_aligned,
            'po': po_aligned,
            'weeks': all_weeks,
            'week_labels': week_labels
        }

    def calculate_top_gaps(self, gap_df, percentile=70):
        """
        计算top N%的差异值

        Args:
            gap_df: 差异DataFrame
            percentile: 百分位数（默认70，即top30%）

        Returns:
            float: 阈值
        """
        # 获取所有差异的绝对值
        abs_gaps = gap_df.abs().values.flatten()

        # 去除0值
        abs_gaps = abs_gaps[abs_gaps > 0]

        if len(abs_gaps) == 0:
            return 0

        # 计算阈值（top30%）
        threshold = np.percentile(abs_gaps, percentile)

        return threshold

    def export_to_excel(self, output_path, highlight_top_percent=30):
        """
        导出差异表到Excel（按周维度），并高亮显示top N%的差异

        Args:
            output_path: 输出文件路径
            highlight_top_percent: 高亮百分比（默认30%）
        """
        # 生成差异数据（按周）
        data = self.create_gap_table()
        gap_df = data['gap']
        schedule_df = data['schedule']
        po_df = data['po']
        weeks = data['weeks']
        week_labels = data['week_labels']

        # 计算高亮阈值
        threshold = self.calculate_top_gaps(gap_df, 100 - highlight_top_percent)

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "差异分析"

        # 设置列宽
        ws.column_dimensions['A'].width = 15

        # 定义样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        subheader_font = Font(bold=True, size=10)

        highlight_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        normal_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )

        # 写入数据
        current_col = 1
        current_row = 1

        # 1. 写入第一行分类标题
        ws.cell(row=1, column=1, value="SKU")
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

        current_col = 2

        # GAP差异部分
        gap_start_col = current_col
        ws.cell(row=1, column=current_col, value="GAP差异")
        ws.cell(row=1, column=current_col).fill = header_fill
        ws.cell(row=1, column=current_col).font = header_font
        ws.cell(row=1, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
        current_col += len(weeks)

        # 排程目标部分
        schedule_start_col = current_col
        ws.cell(row=1, column=current_col, value="排程目标")
        ws.cell(row=1, column=current_col).fill = header_fill
        ws.cell(row=1, column=current_col).font = header_font
        ws.cell(row=1, column=current_col).alignment = Alignment(horizontal='center', vertical='center')
        current_col += len(weeks)

        # PO汇总结果部分
        po_start_col = current_col
        ws.cell(row=1, column=current_col, value="PO汇总结果")
        ws.cell(row=1, column=current_col).fill = header_fill
        ws.cell(row=1, column=current_col).font = header_font
        ws.cell(row=1, column=current_col).alignment = Alignment(horizontal='center', vertical='center')

        # 合并分类标题单元格
        ws.merge_cells(start_row=1, start_column=gap_start_col,
                      end_row=1, end_column=gap_start_col + len(weeks) - 1)
        ws.merge_cells(start_row=1, start_column=schedule_start_col,
                      end_row=1, end_column=schedule_start_col + len(weeks) - 1)
        ws.merge_cells(start_row=1, start_column=po_start_col,
                      end_row=1, end_column=po_start_col + len(weeks) - 1)

        # 2. 写入第二行周次标题
        ws.cell(row=2, column=1, value="SKU")
        ws.cell(row=2, column=1).fill = subheader_fill
        ws.cell(row=2, column=1).font = subheader_font
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # 写入3遍周次标签（GAP、排程、PO）
        for part_idx in range(3):
            start_col = [gap_start_col, schedule_start_col, po_start_col][part_idx]
            for week_idx, week_label in enumerate(week_labels):
                col = start_col + week_idx
                ws.cell(row=2, column=col, value=week_label)
                ws.cell(row=2, column=col).fill = subheader_fill
                ws.cell(row=2, column=col).font = subheader_font
                ws.cell(row=2, column=col).alignment = Alignment(horizontal='center', vertical='center', text_rotation=45)
                ws.column_dimensions[ws.cell(row=2, column=col).column_letter].width = 10

        # 3. 写入数据行
        for sku_idx, sku in enumerate(gap_df.index):
            row = sku_idx + 3

            # SKU列
            ws.cell(row=row, column=1, value=sku)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row, column=1).border = normal_border

            # GAP差异数据
            for week_idx, week_num in enumerate(weeks):
                col = gap_start_col + week_idx
                value = gap_df.loc[sku, week_num]
                cell = ws.cell(row=row, column=col, value=float(value))
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = normal_border

                # 高亮top30%差异
                if abs(value) >= threshold and abs(value) > 0:
                    cell.fill = highlight_fill
                    cell.font = Font(bold=True)

            # 排程目标数据
            for week_idx, week_num in enumerate(weeks):
                col = schedule_start_col + week_idx
                value = schedule_df.loc[sku, week_num]
                cell = ws.cell(row=row, column=col, value=float(value))
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = normal_border

            # PO汇总结果数据
            for week_idx, week_num in enumerate(weeks):
                col = po_start_col + week_idx
                value = po_df.loc[sku, week_num]
                cell = ws.cell(row=row, column=col, value=float(value))
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = normal_border

        # 冻结首行首列
        ws.freeze_panes = 'B3'

        # 保存文件
        wb.save(output_path)
        print(f"差异分析表已保存到: {output_path}")

        return output_path

    def generate_summary_stats(self):
        """
        生成差异汇总统计

        Returns:
            dict: 统计信息
        """
        data = self.create_gap_table()
        gap_df = data['gap']

        # 计算统计指标
        total_gap = gap_df.sum().sum()
        abs_total_gap = gap_df.abs().sum().sum()
        max_gap = gap_df.max().max()
        min_gap = gap_df.min().min()
        avg_gap = gap_df.mean().mean()

        # 正负差异统计
        positive_gaps = gap_df[gap_df > 0].sum().sum()
        negative_gaps = gap_df[gap_df < 0].sum().sum()

        return {
            'total_gap': float(total_gap),
            'abs_total_gap': float(abs_total_gap),
            'max_gap': float(max_gap),
            'min_gap': float(min_gap),
            'avg_gap': float(avg_gap),
            'positive_gaps': float(positive_gaps),
            'negative_gaps': float(negative_gaps),
            'sku_count': len(gap_df),
            'week_count': len(gap_df.columns)  # 改为周数统计
        }


def main():
    """测试函数"""
    import sys

    if len(sys.argv) < 4:
        print("用法: python gap_analysis.py <schedule_file> <po_original> <po_optimized> [output]")
        return

    schedule_file = sys.argv[1]
    po_original = sys.argv[2]
    po_optimized = sys.argv[3]
    output = sys.argv[4] if len(sys.argv) > 4 else 'gap_analysis.xlsx'

    analyzer = GapAnalyzer(schedule_file, po_original, po_optimized)
    analyzer.export_to_excel(output)

    stats = analyzer.generate_summary_stats()
    print("\n差异统计:")
    for key, value in stats.items():
        print(f"  {key}: {value}")


if __name__ == '__main__':
    main()
